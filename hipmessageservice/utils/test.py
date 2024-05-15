#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""=================================================
    @Project: django_hip_service
    @File： test.py
    @Author：liaozhimingandy
    @Email: liaozhimingandy@gmail.com
    @Date：2024-04-23 16:06
    @Desc: 
================================================="""
import os
import uuid
from datetime import datetime

import pymssql
from lxml import etree
from pypinyin import lazy_pinyin
from openpyxl import load_workbook


def fun1(file_path=r'D:\BaiduNetdiskWorkspace\互联互通服务\services\samples\services\BloodTransAppInfoAdd.xml'):
    # 从文件加载XML数据
    tree = etree.parse(file_path)

    # 使用XPath表达式获取所有属性值
    attrs = tree.xpath("//@codeSystem")

    # 打印数据
    for attr in attrs:
        item = tree.xpath(f"//*[@codeSystem='{attr}']")[0].get('codeSystemName')
        if item:
            print(f"字典表: {item} -> {attrs.index(attr) + 1}")
        else:
            try:
                item = tree.xpath(f"//*[@codeSystem='{attr}']")[0].xpath("./*")[0].get('value')
            except IndexError:
                el = tree.xpath(f"//*[@codeSystem='{attr}']")
                print(f"其他情况: {el} -> {attrs.index(attr) + 1}")
            else:
                print(f"未匹配到字典表,请参考值域内容: {item} -> {attrs.index(attr) + 1}")


def fun2():
    sample_path = r"D:\BaiduNetdiskWorkspace\互联互通服务\services\samples\services"

    files = os.listdir(sample_path)

    for file in files:
        base_name, extension = os.path.splitext(file)  # 拆分文件名和后缀名
        print(f"-----{base_name}-----")
        file_path = os.path.join(sample_path, file)
        fun1(file_path)


def get_initials(chinese_text):
    # 使用lazy_pinyin获取拼音，参数style设置为NORMAL表示返回不带声调的拼音
    pinyin_list = lazy_pinyin(chinese_text)
    # 提取每个拼音的首字母
    initials = [pinyin[0].upper() for pinyin in pinyin_list if pinyin]
    return ''.join(initials).replace(' ', '')


def db(sql: str, database: str = 'MDMDB'):
    # 连接参数
    server = '10.100.1.142'
    user = 'caradigm'
    password = 'Admin@123'
    database = 'MDMDB'
    # 创建连接
    conn = pymssql.connect(server, user, password, database)
    cursor = conn.cursor(as_dict=True)  # as_dict=True 返回字典而不是元组

    cursor.execute(sql)

    cursor.close()
    conn.commit()


def fun3():
    file_path = r"D:\data\WeChat\WeChat Files\wxid_oviup54trb2i21\FileStorage\File\2024-04\科室.xlsx"
    import pymssql
    import uuid
    from datetime import datetime
    # 连接参数
    server = '10.100.1.142'
    user = 'caradigm'
    password = 'Admin@123'
    database = 'MDMDB'

    # 创建连接
    conn = pymssql.connect(server, user, password, database)
    cursor = conn.cursor(as_dict=True)  # as_dict=True 返回字典而不是元组

    # 加载Excel工作簿
    wb = load_workbook(filename=file_path)

    # 选择活动工作表，或者通过名称/索引选择其他工作表
    sheet = wb.worksheets[0]  # 默认选择活动工作表

    index = 1
    # 遍历行和列
    for row in sheet.iter_rows():
        if index == 1:
            index += 1
            continue
        sql = f"""insert into md_department(MBR_UID, CODE, NAME, PY_CODE, MBR_VER_NO, SORT_NO, IS_ENABLED, DELETE_FLAG, IS_PUBLISH,
                           CREATED_DATE_TIME, CREATED_USER_ID, CREATED_USER, UPDATED_DATE_TIME, UPDATED_USER_ID,
                           UPDATED_USER, DELETED_DATE_TIME, DELETED_USER_ID, DELETED_USER) 
                           values('{uuid.uuid4()}','{row[0].value}', '{row[1].value}', '{get_initials(row[1].value)}', 1, {index}, 1,  0, 2,
                           '{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 1, 1, '{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 1,
                           1, null, null, null)"""
        print(sql)
        cursor.execute(sql)
        index += 1

    cursor.close()
    conn.commit()


def fun4():
    data = "EMPLOYEE_ID$员工主索引@PK_EMPLOYEE$员工主键@EMPLOYE_CODE$员工工号@EMPLOYE_NAME$员工姓名@DATE_OF_BIRTH$出生日期@ID_TYPE_CODE$证件类型代码@ID_TYPE_NAME$证件类型名称@ID_NO$证件号@NATIONALITY_CODE$国籍代码@NATIONALITY_NAME$国籍名称@ETHNIC_CODE$民族代码@ETHNIC_NAME$民族名称@MARITAL_STATUS_CODE$婚姻状况代码@MARITAL_STATUS_NAME$婚姻状况名称@EDUCATION_CODE$学历代码@EDUCATION_NAME$学历名称@DEGREE_CODE$学位代码@DEGREE_NAME$学位名称@POLITICAL_CODE$政治面貌代码@POLITICAL_NAME$政治面貌名称@BIRTH_PLACE_PROVINCE$出生地-省(自治区、直辖市)@BIRTH_PLACE_CITY$出生地-市(地区、州)@BIRTH_PLACE_COUNTY$出生地-县(区)@BIRTH_PLACE$出生地@NATIVE_PLACE_PROVINCE$籍贯-省(自治区、直辖市)@NATIVE_PLACE_CITY$籍贯-市(地区、州)@NATIVE_PLACE$籍贯@HOME_PHONE_NO$家庭电话@PRESENT_ADDRESS_PROVINCE$现住址-省(自治区、直辖市)@PRESENT_ADDRESS_CITY$现住址-市(地区、州)@PRESENT_ADDRESS_COUNTY$现住址-县(区)@PRESENT_ADDRESS_COUNTRY$现住址-乡(镇、街道办事处)@PRESENT_ADDRESS_VILLAGE$现住址-村(街、路、弄等)@PRESENT_ADDRESS_HOUSE_NO$现住址-门牌号码@PRESENT_ADDRESS$现住址@PRESENT_ADDRESS_POSTAL_CODE$现住址（邮编）@PHONE_NO$联系电话@WORK_PHONE_NO$工作单位电话@TITLE_CODE$职称代码@TITLE_NAME$职称名称@JOB_CODE$职务代码@JOB_NAME$职务名称@POSITION_CODE$职位代码（岗位）@POSITION_NAME$职位名称（岗位）@POSITION_TYPE_CODE$职位类别代码@POSITION_TYPE_NAME$职位类别名称@POSTTION_CATEGORY_CODE$岗位类别代码@POSTTION_CATEGORY_NAME$岗位类别名称@POSTTION_GRADE_CODE$岗位等级代码@POSTTION_GRADE_NAME$岗位等级名称@EMAIL$邮箱@EMPLOYE_CLASS_CODE$用人形式代码@EMPLOYE_CLASS_NAME$用人形式名称@$科室名称@SORT_NO$排序码@SPELL_CODE$拼音简码@WB_CODE$五笔简码@CUSTOM_CODE$自定义简码@CMMT$备注说明@VALID_STATE$有效性标志@ORG_CODE$机构/院部代码@ORG_NAME$机构/院部名称@SOURCE_SYSTEM_CODE$源系统代码@ENTER_OPERA_ID$录入人ID@ENTER_OPERA_NAME$录入人姓名@ENTER_DATE_TIME$录入日期时间@MODIFY_OPERA_ID$修改人ID@MODIFY_OPERA_NAME$修改人姓名@MODIFY_DATE_TIME$修改日期@EXPERTISE$个人专长@INTRODUCTION$个人简介"

    index = 9
    for el in data.split('@'):
        item = el.split('$')
        print(f"('{uuid.uuid4()}','6CB7679592FA47EA828FBC5081BFD309' ,'{item[0]}' ,'{item[1]}','{item[0]}' "
              f",'{item[0]}'  ,0 , 0 ,0 , {index} ,1 ,0 ,1 ,1, '{get_initials(item[1])}', 'STRING',100,0),")
        index += 1


def fun5():
    file_path = r'C:\Users\liaoz\JCPT_EMPLOYEE.csv'
    import csv
    with open(file_path, mode='r') as file:
        csv_reader = csv.DictReader(file)
        data = list(csv_reader)

    index = 1
    for item in data:
        index += 1
        del item['SORT_NO']
        keys = list(item.keys())

        content = "', '".join([item[key] for key in keys]) + "'"
        content2 = f"""'{uuid.uuid4()}','{item['EMPLOYE_CODE']}', '{item['EMPLOYE_NAME']}', '{get_initials(item['EMPLOYE_NAME'])}', 1, {index}, 1,  0, 2, '{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 1, 1, '{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 1,1, null, null, null"""
        sql = f"""insert into [MD_employee](EMPLOYEE_ID, PK_EMPLOYEE, EMPLOYE_CODE, EMPLOYE_NAME, GENDER_CODE, GENDER_NAME, DATE_OF_BIRTH, ID_TYPE_CODE, ID_TYPE_NAME, ID_NO, NATIONALITY_CODE, NATIONALITY_NAME, ETHNIC_CODE, ETHNIC_NAME, MARITAL_STATUS_CODE, MARITAL_STATUS_NAME, EDUCATION_CODE, EDUCATION_NAME, DEGREE_CODE, DEGREE_NAME, POLITICAL_CODE, POLITICAL_NAME, BIRTH_PLACE_PROVINCE, BIRTH_PLACE_CITY, BIRTH_PLACE_COUNTY, BIRTH_PLACE, NATIVE_PLACE_PROVINCE, NATIVE_PLACE_CITY, NATIVE_PLACE, HOME_PHONE_NO, PRESENT_ADDRESS_PROVINCE, PRESENT_ADDRESS_CITY, PRESENT_ADDRESS_COUNTY, PRESENT_ADDRESS_COUNTRY, PRESENT_ADDRESS_VILLAGE, PRESENT_ADDRESS_HOUSE_NO, PRESENT_ADDRESS, PRESENT_ADDRESS_POSTAL_CODE, PHONE_NO, WORK_PHONE_NO, TITLE_CODE, TITLE_NAME, JOB_CODE, JOB_NAME, POSITION_CODE, POSITION_NAME, POSITION_TYPE_CODE, POSITION_TYPE_NAME, POSTTION_CATEGORY_CODE, POSTTION_CATEGORY_NAME, POSTTION_GRADE_CODE, POSTTION_GRADE_NAME, EMAIL, EMPLOYE_CLASS_CODE, EMPLOYE_CLASS_NAME, DEPT_ID, DEPT_NAME, SPELL_CODE, WB_CODE, CUSTOM_CODE, CMMT, VALID_STATE, ORG_CODE, ORG_NAME, SOURCE_SYSTEM_CODE, ENTER_OPERA_ID, ENTER_OPERA_NAME, ENTER_DATE_TIME, MODIFY_OPERA_ID, MODIFY_OPERA_NAME, MODIFY_DATE_TIME, EXPERTISE, INTRODUCTION,MBR_UID,CODE,[NAME],PY_CODE,MBR_VER_NO,SORT_NO,IS_ENABLED,DELETE_FLAG,IS_PUBLISH,CREATED_DATE_TIME,CREATED_USER_ID,CREATED_USER,UPDATED_DATE_TIME,UPDATED_USER_ID,UPDATED_USER,DELETED_DATE_TIME,DELETED_USER_ID,DELETED_USER)
                values('{content},{content2})"""
        db(sql)
    print(f"导入完成;共计{index}")


def fun6(file_path: str, table_name: str = 'md_department', sheet_name: str = '国籍'):

    # 加载Excel工作簿
    wb = load_workbook(filename=file_path)

    # 选择活动工作表，或者通过名称/索引选择其他工作表
    sheet = wb[sheet_name] # 默认选择活动工作表

    index = 0
    # 遍历行和列
    for row in sheet.iter_rows():
        index += 1
        if index < 8:
            continue
        sql = f"""insert into {table_name}(MBR_UID, CODE, NAME, PY_CODE, MBR_VER_NO, SORT_NO, IS_ENABLED, DELETE_FLAG, IS_PUBLISH,
                               CREATED_DATE_TIME, CREATED_USER_ID, CREATED_USER, UPDATED_DATE_TIME, UPDATED_USER_ID,
                               UPDATED_USER, DELETED_DATE_TIME, DELETED_USER_ID, DELETED_USER) 
                               values('{uuid.uuid4()}','{row[0].value}', '{row[1].value.strip()}', '{get_initials(row[1].value.strip())}', 1, {index}, 1,  0, 2,
                               '{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 1, 1, '{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 1,
                               1, null, null, null)"""
        print(sql)
        # db(sql)


if __name__ == "__main__":
    file_path = r"C:\Users\liaoz\Downloads\萨米数据标准-字典标准梳理修订20240328.xlsx"
    fun6(file_path, table_name='MD_rh_blood_type', sheet_name="RH血型")
