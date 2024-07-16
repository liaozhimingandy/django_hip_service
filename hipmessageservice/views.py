import base64
import datetime
from http import HTTPStatus

import lxml
import openpyxl
import pandas as pd
import requests
from django.conf import settings
from django.contrib.staticfiles import finders
from django.db import connection
from django.db import transaction
from django.http import FileResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.cache import never_cache
from json2xml import json2xml
from lxml import etree
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.styles.fills import PatternFill

from cdr.models import Patient
# from .authentication import AuthBearer
from hipmessageservice.models import Service, Application, StatusShip
from hipmessageservice.utils.database import read_cda


# Create your views here.
@never_cache
@transaction.non_atomic_requests
def index(request):
    services = (Service.objects.filter(is_deleted=False)
                .only('service_id', 'service_name', 'service_code', 'is_v3', "id", "service_queue")).order_by(
        "service_queue")
    status = (
        StatusShip.objects.filter(is_deleted=False, service__is_deleted=False).only("service_id", "application_id",
                                                                                    "status"))
    dict_status = {}
    for obj in status:
        dict_status[f"{obj.service_id}-{obj.application_id}"] = obj.status

    applications = (Application.objects.filter(is_deleted=False, firm__isnull=False).order_by('-firm')
                    .prefetch_related('firm').only("id", 'application_name', "firm_id", 'firm__firm_name_short'))

    return render(request, 'hipmessageservice/index.html',
                  context={'services': services, 'applications': applications, 'status': status,
                           'dict_status': dict_status})


def read_all() -> list:
    sql = """select a.service_code, a.service_name, d.firm_name_short, c.application_name, b.status, a.service_queue
            from hipmessageservice_service a
            right join hipmessageservice_statusship b on a.id = b.service_id and b.is_deleted=False
            right join hipmessageservice_application c on c.id = b.application_id and b.is_deleted=False
            right join hipmessageservice_firm d on d.id = c.firm_id
            where a.is_deleted = False
            order by a.service_queue asc;
            """

    with connection.cursor() as cursor:
        cursor.execute(sql)
        items = cursor.fetchall()

    return list(items)


def generate_cda(df_all):
    """
    生成cda分工sheet
    :return:
    """

    list_cda = read_cda()
    str_sheet_name = '分工-CDA'

    df = (pd.DataFrame(list_cda, columns=['value', 'comment', 'firm__firm_name_short', "count"]))
    result = df.pivot_table(values="count", index=['value', 'comment'], columns=["firm__firm_name_short"])

    # 替换值
    result.replace(1, '√', inplace=True)
    writer = pd.ExcelWriter("temp/results.xlsx", engine="xlsxwriter")

    # 保存交互场景
    df_all.to_excel(writer, sheet_name="交互场景", index=True, index_label=0)

    # 设置样式
    format_sheet = writer.book.add_format({'fg_color': '#D7E4BC', 'bold': True, 'align': 'center', 'valign': 'vcenter',
                                           'border': 2, 'text_wrap': True})
    result.to_excel(writer, index=True, sheet_name=str_sheet_name, index_label=['CDA文档代码', 'CDA文档名称'])
    sheet = writer.sheets[str_sheet_name]
    sheet.set_column('A1:I47', 18, cell_format=format_sheet)

    writer.close()


def cell_style(sheet, cell, style_name, style):
    sheet[cell][style_name] = style


def generate_data():
    map_status = {1: '订阅', 2: "发布", 3: "全部"}
    filename = "temp/results.xlsx"
    statuss = StatusShip.objects.filter(is_deleted=False).only("application__firm__firm_name_short").distinct(
        "application__firm__firm_name_short")

    # 样式
    font = Font(size=16, bold=True, color='FF000000')
    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'))
    orange_fill = PatternFill(fill_type='solid', fgColor="1890ff")
    # 居中所有单元格
    align = Alignment(horizontal='center', vertical='center', wrapText=False)

    for status in statuss:
        if not status.application.firm:
            continue

        book = openpyxl.load_workbook(filename)
        sheet = book.create_sheet(title=f"交互服务分工-{status.application.firm.firm_name_short}")
        index = 1

        items = (StatusShip.objects.filter(application__firm__firm_name_short=status.application.firm.firm_name_short,
                                           service__is_deleted=False).values('service__service_name',
                                                                             'service__service_code', 'status')
                 .distinct('service__service_name', 'service__service_code', 'status'))

        sheet[f'A{index}'] = f"以下为您需要完成的交互服务|统计日期:{datetime.datetime.now().strftime('%Y-%m-%d')}"
        # cell_style(sheet, f'A{index}', "font", font)
        sheet[f'A{index}'].font = font
        sheet[f'A{index}'].border = border
        sheet[f'A{index}'].alignment = align
        sheet.merge_cells('A1:C1')

        index += 1

        sheet[f'A{index}'] = "服务名称"
        sheet[f'B{index}'] = "服务代码"
        sheet[f'C{index}'] = "发布订阅关系"
        sheet[f'A{index}'].font = font
        sheet[f'B{index}'].font = font
        sheet[f'C{index}'].font = font
        sheet[f'A{index}'].border = border
        sheet[f'B{index}'].border = border
        sheet[f'C{index}'].border = border
        sheet[f'A{index}'].fill = orange_fill
        sheet[f'B{index}'].fill = orange_fill
        sheet[f'C{index}'].fill = orange_fill
        sheet[f'A{index}'].alignment = align
        sheet[f'B{index}'].alignment = align
        sheet[f'C{index}'].alignment = align

        for item in items:
            sheet[f'A{index + 1}'] = item['service__service_name']
            sheet[f'B{index + 1}'] = item['service__service_code']
            sheet[f'C{index + 1}'] = map_status.get(item['status'])

            sheet[f'A{index + 1}'].font = font
            sheet[f'B{index + 1}'].font = font
            sheet[f'C{index + 1}'].font = font
            sheet[f'A{index + 1}'].border = border
            sheet[f'B{index + 1}'].border = border
            sheet[f'C{index + 1}'].border = border

            index += 1

        sheet.column_dimensions["A"].width = 60
        sheet.column_dimensions["B"].width = 60
        sheet.column_dimensions["C"].width = 60
        book.save(filename)

        # break


def generate_service_all():
    list_data = read_all()
    df = pd.DataFrame.from_records(list_data, columns=['service_code', 'service_name', 'firm_name_short',
                                                       'application_name', 'status', 'service_queue'])

    result = df.pivot_table(values="status", index=['service_queue', 'service_code', 'service_name'],
                            columns=["firm_name_short", 'application_name'])  # 暂不支持dropna=False

    result.sort_values(by='service_queue', ascending=True)
    result.replace(1, '订阅', inplace=True)
    result.replace(2, '发布', inplace=True)
    result.replace(3, '全部', inplace=True)

    # result.to_excel("temp/results.xlsx", sheet_name='交互场景', index=True)
    return result


def generate_report(request):
    generate_service_all()
    return HttpResponse("ok", content_type='text')


def download(request):
    import os
    if not os.path.exists("temp"):
        os.makedirs("temp")
    if os.path.exists("temp/results.xlsx"):
        os.remove("temp/results.xlsx")

    # 生成消息场景
    df = generate_service_all()
    # 生成cda统计数据
    generate_cda(df)
    # 生成交互服务分工
    generate_data()

    return FileResponse(open('temp/results.xlsx', 'rb'), as_attachment=True,
                        filename=f'医院信息平台交互规范-交互场景-{datetime.datetime.now().strftime("%Y-%m-%d")}.xlsx')


def home(request):
    services = (Service.objects.filter(is_deleted=False)
                .only('service_id', 'service_name', 'service_code', 'is_v3', "id"))
    status = (
        StatusShip.objects.filter(is_deleted=False, service__is_deleted=False).only("service_id", "application_id",
                                                                                    "status"))
    dict_status = {}
    for obj in status:
        dict_status[f"{obj.service_id}-{obj.application_id}"] = obj.status

    applications = (Application.objects.filter(is_deleted=False, firm__isnull=False).order_by('-firm')
                    .prefetch_related('firm').only("id", 'application_name', "firm_id", 'firm__firm_name_short'))

    return render(request, 'hipmessageservice/home.html',
                  context={'services': services, 'applications': applications, 'status': status,
                           'dict_status': dict_status, 'user': request.user})


def test3(request):
    numbers = range(1, 100)
    return render(request, 'hipmessageservice/test3.html',
                  context={'user': request.user, 'numbers': numbers})


data_mapping = {
    "patient_id": "//xmlns:patient/xmlns:id/xmlns:item/@extension",
    "gmt_reg": "//xmlns:patient/xmlns:effectiveTime/xmlns:any/@value",
    "id_no": "//xmlns:patientPerson/xmlns:id/xmlns:item/@extension",
    "id_code": "//xmlns:patientPerson/xmlns:idCategory/@code",
    "id_name": "//xmlns:patientPerson/xmlns:idCategory/xmlns:displayName/@value",
    "patient_name": "//xmlns:patientPerson/xmlns:name/xmlns:item/xmlns:part/@value",
    "tel_no": "//xmlns:patientPerson/xmlns:telecom/xmlns:item/@value",
    "sex_code": "//xmlns:administrativeGenderCode/@code",
    "sex_name": "//xmlns:administrativeGenderCode/xmlns:displayName/@value",
    "gmt_birth": "//xmlns:birthTime/@value",
    "addr_sal": "//xmlns:addr/xmlns:item/xmlns:part[@type='SAL']/@value",
    "addr_sta": "//xmlns:addr/xmlns:item/xmlns:part[@type='STA']/@value",
    "addr_cty": "//xmlns:addr/xmlns:item/xmlns:part[@type='CTY']/@value",
    "addr_cnt": "//xmlns:addr/xmlns:item/xmlns:part[@type='CNT']/@value",
    "addr_stb": "//xmlns:addr/xmlns:item/xmlns:part[@type='STB']/@value",
    "addr_str": "//xmlns:addr/xmlns:item/xmlns:part[@type='STR']/@value",
    "addr_bnr": "//xmlns:addr/xmlns:item/xmlns:part[@type='BNR']/@value",
    "addr_zip": "//xmlns:addr/xmlns:item/xmlns:part[@type='ZIP']/@value",
    "marital_status_code": "//xmlns:maritalStatusCode/@code",
    "marital_status_name": "//xmlns:maritalStatusCode/xmlns:displayName/@value",
    "occupation_code": "//xmlns:occupationCode/@code",
    "occupation_name": "//xmlns:occupationCode/xmlns:displayName/@value",
    "ethnic_group_code": "//xmlns:ethnicGroupCode/xmlns:item/xmlns:displayName/@value",
    "ethnic_group_name": "//xmlns:ethnicGroupCode/xmlns:item/@code",
    "work_org": "//xmlns:employerOrganization/xmlns:name/xmlns:item/xmlns:part/@value",
    "work_org_tel": "//xmlns:employerOrganization/xmlns:contactParty/xmlns:telecom/xmlns:item/@value",
    "hcard_no": "//xmlns:asOtherIDs/xmlns:id/xmlns:item[@root='2.16.156.10011.1.19']/@extension",
    "hcard_org_code": "//xmlns:asOtherIDs/xmlns:id/xmlns:item["
                      "@root='2.16.156.10011.1.19']/parent::*/following-sibling::*/xmlns:id/xmlns:item/@extension",
    "gcard_no": "//xmlns:asOtherIDs/xmlns:id/xmlns:item[@root='2.16.156.10011.1.2']/@extension",
    "gcard_org_code": "//xmlns:asOtherIDs/xmlns:id/xmlns:item["
                      "@root='2.16.156.10011.1.2']/parent::*/following-sibling::*/xmlns:id/xmlns:item/@extension",
    "contact_code": "//xmlns:personalRelationship/xmlns:code/@code",
    "contact_name": "//xmlns:personalRelationship/xmlns:code/xmlns:displayName/@value",
    "contact_cname": "//xmlns:relationshipHolder1/xmlns:name/xmlns:item/xmlns:part/@value",
    "contact_tel": "//xmlns:personalRelationship/xmlns:telecom/xmlns:item/@value",
    "org_code": "//xmlns:providerOrganization/xmlns:name/xmlns:item/xmlns:part/@value",
    "org_name": "//xmlns:providerOrganization/xmlns:id/xmlns:item/@extension",
    "ins_code": "//xmlns:beneficiary/xmlns:code/@code",
    "ins_name": "//xmlns:beneficiary/xmlns:code/xmlns:displayName/@value",
    "author_id": "//xmlns:author/xmlns:assignedEntity/xmlns:assignedPerson/xmlns:name/xmlns:item/xmlns:part/@value",
    "author": "//xmlns:author/xmlns:assignedEntity/xmlns:id/xmlns:item/@extension",
    "from_src": "//xmlns:sender/xmlns:device/xmlns:id/xmlns:item/@extension"
}

dict_empi_mapping = {
    "SYS_REC_ID": "patient_id",
    "LAST_UPDATE_TIME": "gmt_reg",
    "ID_NO": "id_no",
    "ID_TYPE_CODE": "id_code",
    "ID_TYPE_NAME": "id_name",
    "NAME": "patient_name",
    "PHONE_NO": "tel_no",
    "GENDER_CODE": "sex_code",
    "GENDER_NAME": "sex_name",
    "DOB": "gmt_birth",
    "PRESENT_ADDRESS_PROVINCE": "addr_sta",
    "PRESENT_ADDRESS_CITY": "addr_cty",
    "PRESENT_ADDRESS_COUNTY": "addr_cnt",
    "PRESENT_ADDRES_COUNTRY": "addr_stb",
    "PRESENT_ADDRES_VILLAGE": "addr_str",
    "PRESENT_ADDRES_HOUSE_NO": "addr_bnr",
    "PRESENT_ADDRES_POSTAL_CODE": "addr_zip",
    "MARITAL_STATUS_CODE": "marital_status_code",
    "MARITAL_STATUS_NAME": "marital_status_name",
    "ETHNIC_GROUP_CODE": "ethnic_group_code",
    "ETHNIC_GROUP_NAME": "ethnic_group_name",
    "OCCUPATION_CODE": "occupation_code",
    "OCCUPATION_NAME": "occupation_name",
    "EMPLOYER_NAME": "work_org",
    "EMPLOYER_PHONE_NO": "work_org_tel",
    "HEALTH_CARD_NO": "hcard_no",
    "HEALTH_RECORD_NO": "gcard_no",
    "CONTACT_RELATIONSHIP_CODE": "contact_code",
    "CONTACT_RELATIONSHIP_NAME": "contact_name",
    "CONTACT_PHONE_NO": "contact_tel",
    "CONTACT_NAME": "contact_cname",
    "ORG_CODE": "org_code",
    "MEDICAL_INSURANCE_TYPE_CODE": "ins_code",
    "MEDICAL_INSURANCE_TYPE_NAME": "ins_name",
    "LAST_UPDATE_BY": "author_id",
    "SYS_CODE": "from_src",
    "MEDICAL_INSURANCE_CARD_NO": "",
    "CITIZENSHIP_CODE": "",
    "CITIZENSHIP_NAME": "",
    "EDUCATION_LEVEL_CODE": "",
    "EDUCATION_LEVEL_NAME": "",
    "POB_PROVINCE": "",
    "POB_CITY": "",
    "POB_COUNTY": "",
    "NATIVE_PLACE_PROVINCE": "",
    "NATIVE_PLACE_CITY": "",
    "EMPLOYER_PROVINCE": "",
    "EMPLOYER_CITY": "",
    "EMPLOYER_COUNTY": "",
    "EMPLOYER_POSTAL_CODE": "",
    "CONTACT_ADDRESS": "",
    "ABO_BLOOD_TYPE_CODE": "",
    "ABO_BLOOD_TYPE_NAME": "",
    "RH_BLOOD_TYPE_CODE": "",
    "RH_BLOOD_TYPE_NAME": "",
    "CREATE_BY": "",
    "CREATE_TIME": "",
    "MEDICAL_RECORD_NO": "",
    "OP_MEDICAL_RECORD_NO": "",
    "Passport_GA": "",
    "BEIJING_MEDICAL_CARD_NO": "",
    "VIP_INDICATOR": "",
    "DEATH_INDICATOR": "",
    "DOD": "",
    "HOSPITAL_CARD_NO": "",
    "CATEGORY_CODE": ""
}


def DealPatient(content):
    """处理个人新增注册,更新接口,调取empi注册接口,并且保存到数据库"""
    from lxml import etree
    etree.register_namespace('xmlns', 'urn:hl7-org:v3')
    xml_root = etree.fromstring(content)
    dict_data = data_mapping.copy()

    # 解析平台患者个人信息数据集所需要的数据
    for key in data_mapping.keys():
        data = xml_root.xpath(data_mapping[key], namespaces={'xmlns': 'urn:hl7-org:v3'})
        dict_data[key] = data[0]

    # 日期处理
    dict_data['gmt_reg'] = timezone.make_aware(datetime.datetime.strptime(dict_data['gmt_reg'], '%Y%m%d%H%M%S'),
                                               timezone.get_current_timezone())
    dict_data['gmt_birth'] = timezone.make_aware(datetime.datetime.strptime(dict_data['gmt_birth'], '%Y%m%d'),
                                                 timezone.get_current_timezone())
    dict_data_empi = {}

    # 调empi接口
    for key in dict_empi_mapping.keys():
        if dict_empi_mapping[key]:
            dict_data_empi[key] = dict_data[dict_empi_mapping[key]]

    xml_data = json2xml.Json2xml(dict_data_empi, pretty=True, wrapper="EMPI_PERSON").to_xml()
    payload = f"""
            <soap:Envelope xmlns:soap="http://www.w3.org/2003/05/soap-envelope" xmlns:car="http://www.caradigm.com/">
           <soap:Header/>
           <soap:Body>
              <car:IndexRegisterFunc>
                 <!--Optional:-->
                 <car:xml>
                 <![CDATA[{xml_data.replace('<?xml version="1.0" ?>', '')}]]>
                 </car:xml>
              </car:IndexRegisterFunc>
           </soap:Body>
        </soap:Envelope>
        """
    response = requests.post(settings.EMPI_API_URL, data=payload, headers={'Content-Type': 'text/xml'}, timeout=2)
    empi = None
    if response.status_code == 200:
        # 匹配empi号
        # empi = (re.search('empi_id&gt;\d+&lt;/empi_id', response.text).group().replace('empi_id&gt;', '')
        #         .replace('&lt;/empi_id', ''))
        root = etree.fromstring(response.content)
        xml_empi = root.find('.//xmlns:IndexRegisterFuncResult', namespaces={"xmlns": "http://www.caradigm.com/"}).text
        empi_root = etree.fromstring(xml_empi)

        if empi_root.find('.//state').text == 'success':
            empi = empi_root.find('.//empi_id').text
        else:
            return False, empi_root.find('.//message').text

    else:
        return False, response.text

    dict_data.update(**{"empi": empi})

    # 保存到数据库
    # 移除不需要的key
    list_to_remove = ['id_name', 'sex_name', 'marital_status_name', 'occupation_name', 'ethnic_group_name',
                      'contact_name', 'ins_name']

    for key in list_to_remove:
        dict_data.pop(key, None)

    # 保存到数据库, 耗时比较长
    PatientInfo.objects.update_or_create(patient_id=dict_data['patient_id'], defaults=dict_data)

    return True, f"#empi:{empi}#"
